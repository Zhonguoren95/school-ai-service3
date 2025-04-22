import pandas as pd
import pdfplumber
import re
from openpyxl import load_workbook
from io import BytesIO
from pdf2image import convert_from_bytes
import pytesseract
from gpt_utils import analyze_position_with_gpt
import streamlit as st

def extract_text_from_pdf(file):
    text = ""
    try:
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        if not text.strip():
            file.seek(0)
            images = convert_from_bytes(file.read(), dpi=300)
            for img in images:
                text += pytesseract.image_to_string(img, lang="rus") + "\n"
        return text
    except Exception as e:
        return f"[ERROR] PDF parse failed: {e}"

def parse_requirements(text):
    rows = []
    lines = text.split("\n")
    for line in lines:
        if any(char.isdigit() for char in line):
            parts = re.split(r"\s{2,}|\t", line.strip())
            if len(parts) >= 2:
                name = parts[0]
                quantity = re.search(r"\d+", parts[1])
                quantity = quantity.group() if quantity else ""
                rows.append({"Наименование из ТЗ": name, "Кол-во": quantity})
    return pd.DataFrame(rows)

def load_price_list(files):
    all_items = []
    for file in files:
        df = pd.read_excel(file, header=None)
        for index, row in df.iterrows():
            for col in row:
                if isinstance(col, str) and any(x in col.lower() for x in ["стол", "кресло", "лампа", "шкаф", "банкетка", "барьер"]):
                    price = next((v for v in row if isinstance(v, (int, float))), "")
                    item = {
                        "Артикул": str(row[0]) if len(row) > 0 else "",
                        "Наименование": col,
                        "Цена": price,
                        "Поставщик": file.name
                    }
                    all_items.append(item)
                    break
    return pd.DataFrame(all_items)

def load_discounts(file):
    try:
        df = pd.read_excel(file)
        discounts = {}
        for _, row in df.iterrows():
            supplier = row.get("Поставщик")
            discount = row.get("Скидка", 0)
            if pd.notna(supplier):
                discounts[supplier] = discount
        return discounts
    except:
        return {}

def process_documents(spec_file, prices_files, discounts_file=None):
    log = []
    try:
        text = extract_text_from_pdf(spec_file)
        if text.startswith("[ERROR]"):
            return text, pd.DataFrame(), None

        log.append("📄 ТЗ успешно распознано")
        ts_df = parse_requirements(text)
        log.append(f"🔍 Найдено позиций в ТЗ: {len(ts_df)}")

        enriched_rows = []
        for i, row in enumerate(ts_df.itertuples(), start=1):
            row_dict = row._asdict()
            try:
                analysis = analyze_position_with_gpt(row_dict["Наименование из ТЗ"])
                row_dict["GPT_тип"] = analysis.get("тип", "")
                row_dict["GPT_синонимы"] = ", ".join(analysis.get("синонимы", []))
                row_dict["GPT_ключи"] = ", ".join(analysis.get("ключи", []))
            except Exception as e:
                row_dict["GPT_тип"] = "Ошибка"
                row_dict["GPT_синонимы"] = ""
                row_dict["GPT_ключи"] = row_dict["Наименование из ТЗ"].split()[0]
                log.append(f"⚠️ GPT ошибка в строке {i}: {e}")
            enriched_rows.append(row_dict)

        enriched_df = pd.DataFrame(enriched_rows)
        log.append("✅ GPT-анализ завершён")

        prices_df = load_price_list(prices_files)
        log.append(f"📦 Прайсы загружены: {len(prices_df)} товаров")
        discounts = load_discounts(discounts_file) if discounts_file else {}

        results = []
        for i, req in enriched_df.iterrows():
            try:
                name = req["Наименование из ТЗ"]
                qty = req["Кол-во"]
                search_keys = req.get("GPT_ключи", name).split(", ")

                matched = pd.DataFrame()
                for key in search_keys:
                    matches = prices_df[prices_df["Наименование"].str.contains(key, case=False, na=False)]
                    matched = pd.concat([matched, matches])
                matched = matched.drop_duplicates().sort_values("Цена").head(3)

                item = {
                    "Наименование из ТЗ": name,
                    "Кол-во": qty
                }

                for j, (_, match) in enumerate(matched.iterrows(), start=1):
                    supplier = match.get("Поставщик", f"Поставщик {j}")
                    price = match.get("Цена")
                    discount = discounts.get(supplier, 0)
                    final_price = round(price * (1 - discount / 100), 2) if price else ""

                    item[f"Поставщик {j}"] = supplier
                    item[f"Цена {j}"] = price
                    item[f"Скидка {j}"] = f"{discount}%"
                    item[f"Итог {j}"] = final_price

                if not matched.empty:
                    results.append(item)
                else:
                    item["Поставщик 1"] = "Не найдено"
                    item["Цена 1"] = ""
                    item["Скидка 1"] = ""
                    item["Итог 1"] = ""
                    results.append(item)
            except Exception as e:
                log.append(f"❌ Ошибка в обработке строки {i + 1}: {e}")

        result_df = pd.DataFrame(results)
        log.append(f"✅ Обработка завершена. Успешно обработано: {len(result_df)} строк")

        wb = load_workbook("Форма для результата.xlsx")
        ws = wb.active
        start_row = 10
        for i, row in result_df.iterrows():
            ws.cell(start_row + i, 1, i + 1)
            ws.cell(start_row + i, 2, row["Наименование из ТЗ"])
            ws.cell(start_row + i, 3, row["Кол-во"])
            ws.cell(start_row + i, 4, row.get("Поставщик 1", ""))
            ws.cell(start_row + i, 5, row.get("Цена 1", ""))
            ws.cell(start_row + i, 6, row.get("Скидка 1", ""))
            ws.cell(start_row + i, 7, row.get("Итог 1", ""))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return "\n".join(log), result_df, output.read()

    except Exception as e:
        return f"[ERROR] Global fail: {e}", pd.DataFrame(), None
